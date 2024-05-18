from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()



ws1 = wb.create_sheet("Mysheet")
count = 1
for row in range(1, 11):
    for col in range(1, 11):
        ws1.cell(row=row, column=col, Value=count)
        count += 1


# Save the file
wb.save("sample.xlsx")